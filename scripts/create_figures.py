import matplotlib.pyplot as plt
from pathlib import Path

import pandas as pd

# =========================================================
# PATHS
# =========================================================

RESULTS_FOLDER = Path("results")

MASTER_RESULTS = RESULTS_FOLDER / "master_results.csv"

FIGURES_FOLDER = RESULTS_FOLDER / "figures"

FIGURES_FOLDER.mkdir(exist_ok=True)

# =========================================================
# LOAD RESULTS
# =========================================================

print()

print("=" * 60)

print("LOADING RESULTS")

print("=" * 60)

results = pd.read_csv(MASTER_RESULTS)

print()

print(f"Rows loaded : {len(results)}")

print(f"Columns     : {len(results.columns)}")

print()

print(results.head())

# =========================================================
# PRONUNCIATION ACCURACY SCORE
# =========================================================

print()

print("Calculating PAS...")

results["PAS"] = results["Similarity"] * 100

print("Done.")

print()

print(results[["Street", "Ground Truth", "PAS"]].head())

# =========================================================
# SAVE UPDATED RESULTS
# =========================================================

output_file = RESULTS_FOLDER / "master_results_with_pas.csv"

results.to_csv(

    output_file,

    index=False

)

# =========================================================
# SAVE EXCEL COPY
# =========================================================

excel_file = RESULTS_FOLDER / "master_results_with_pas.xlsx"

results.to_excel(

    excel_file,

    index=False

)

print()

print("CSV Saved:")

print(output_file)

print()

print("Excel Saved:")

print(excel_file)

print()

print("Saved:")

print(output_file)

# =========================================================
# FIGURE 1
# PAS DISTRIBUTION
# =========================================================

plt.figure(figsize=(8,5))

plt.hist(

    results["PAS"],

    bins=10,

    edgecolor="black"

)

plt.title(

    "Distribution of Pronunciation Accuracy Scores",

    fontsize=14,

    weight="bold"

)

plt.xlabel(

    "Pronunciation Accuracy Score (PAS)",

    fontsize=12

)

plt.ylabel(

    "Number of Comparisons",

    fontsize=12

)

# ---------------------------------------------------------
# IMPROVE SPACING
# ---------------------------------------------------------

plt.subplots_adjust(left=0.35, right=0.95)

plt.xticks(fontsize=10)
plt.yticks(fontsize=10)

plt.tight_layout()
# =========================================================
# FIGURE 2
# DTW DISTRIBUTION
# =========================================================

plt.figure(figsize=(8,5))

plt.hist(

    results["DTW Distance"],

    bins=12,

    edgecolor="black"

)

plt.title(

    "Distribution of DTW Distances",

    fontsize=14,

    weight="bold"

)

plt.xlabel(

    "DTW Distance",

    fontsize=12

)

plt.ylabel(

    "Number of Comparisons",

    fontsize=12

)

plt.tight_layout()

figure_file = FIGURES_FOLDER / "figure02_dtw_distribution.png"

plt.savefig(

    figure_file,

    dpi=300,

    bbox_inches="tight"

)

plt.close()

print()

print("Figure saved:")

print(figure_file)

# =========================================================
# FIGURE 3
# TOP 10 VS BOTTOM 10 PAS RANKING
# =========================================================

import matplotlib.pyplot as plt

# ---------------------------------------------------------
# Average PAS per street
# ---------------------------------------------------------

street_pas = (
    results
    .groupby("Street")["PAS"]
    .mean()
    .sort_values()
)

overall_mean = street_pas.mean()

bottom10 = street_pas.head(10)
top10 = street_pas.tail(10)

ranking = pd.concat([bottom10, top10])

# ---------------------------------------------------------
# Colours
# ---------------------------------------------------------

colours = []

for value in ranking.values:

    if value <= overall_mean * 0.90:
        colours.append("#d62728")      # red

    elif value >= overall_mean * 1.10:
        colours.append("#2ca02c")      # green

    else:
        colours.append("#ffbf00")      # yellow

# ---------------------------------------------------------
# Emoji Labels
# ---------------------------------------------------------

labels = []
label_colours = []

for value in ranking.values:

    if value <= overall_mean * 0.90:

        labels.append("● Lowest")
        label_colours.append("#d62728")

    elif value >= overall_mean * 1.10:

        labels.append("● Highest")
        label_colours.append("#2ca02c")

    else:

        labels.append("● Mid")
        label_colours.append("#ffbf00")

# ---------------------------------------------------------
# Figure
# ---------------------------------------------------------

fig, ax = plt.subplots(figsize=(15,10))

bars = ax.barh(

    ranking.index,

    ranking.values * 100,

    color="#5B84B1",

    edgecolor="black"

)

# ---------------------------------------------------------
# PAS Labels
# ---------------------------------------------------------

for i, bar in enumerate(bars):

    width = bar.get_width()

    y = bar.get_y() + bar.get_height()/2

    # PAS percentage
    ax.text(

        width + 0.5,

        y,

        f"{width:.1f}%",

        va="center",

        fontsize=9,

        fontweight="bold"

    )

    # Coloured traffic-light label
    ax.text(

        width + 8,

        y,

        labels[i],

        color=label_colours[i],

        va="center",

        fontsize=9,

        fontweight="bold"

    )

# ---------------------------------------------------------
# Mean PAS
# ---------------------------------------------------------

ax.axvline(

    overall_mean * 100,

    color="black",

    linestyle="--",

    linewidth=2,

    label=f"Mean PAS ({overall_mean*100:.1f}%)"

)

# ---------------------------------------------------------
# SECTION HEADERS
# ---------------------------------------------------------

# Divider between Bottom 10 and Top 10
ax.axhline(

    9.5,

    color="black",

    linewidth=1.5,

    linestyle="--"

)

# Bottom section title
ax.text(

    -5,

    -1.2,

    "LOWEST PAS STREETS",

    fontsize=13,

    fontweight="bold",

    color="#B22222",

    ha="left"

)

# Top section title
ax.text(

    -5,

    10.2,

    "HIGHEST PAS STREETS",

    fontsize=13,

    fontweight="bold",

    color="#006400",

    ha="left"

)

# Mean PAS annotation
ax.text(

    overall_mean * 100,

    9.9,

    f"Mean PAS = {overall_mean*100:.1f}%",

    fontsize=10,

    fontweight="bold",

    ha="center",

    va="bottom",

    bbox=dict(

        facecolor="white",

        edgecolor="black",

        boxstyle="round,pad=0.3"

    )

)
# ---------------------------------------------------------
# Summary Box
# ---------------------------------------------------------

summary = (

    f"Street Names : {len(street_pas)}\n"

    f"Mean PAS     : {overall_mean*100:.1f}%\n"

    f"Highest PAS  : {street_pas.max()*100:.1f}%\n"

    f"Lowest PAS   : {street_pas.min()*100:.1f}%"

)

ax.text(

    82,

    18,

    summary,

    fontsize=10,

    bbox=dict(

        facecolor="white",

        edgecolor="black",

        boxstyle="round"

    )

)

ax.legend()

plt.tight_layout()

figure_file = FIGURES_FOLDER / "Figure03_PAS_Top10_Bottom10.png"

# ---------------------------------------------------------
# Titles
# ---------------------------------------------------------

ax.set_title(
    "Figure 3. Pronunciation Accuracy Score (PAS) Ranking\n"
    "Comparison of the Lowest and Highest Ranked isiZulu Street Names",
    fontsize=15,
    fontweight="bold",
    pad=20
)

ax.set_xlabel(
    "Average Pronunciation Accuracy Score (%)"
)

ax.set_ylabel(
    "Street Name"
)

# ---------------------------------------------------------
# GRID
# ---------------------------------------------------------

ax.grid(
    axis="x",
    linestyle="--",
    alpha=0.30
)

# ---------------------------------------------------------
# TICK SIZE
# ---------------------------------------------------------

ax.tick_params(
    axis="y",
    labelsize=10
)

ax.tick_params(
    axis="x",
    labelsize=10
)

# ---------------------------------------------------------
# SPACING
# ---------------------------------------------------------

plt.subplots_adjust(
    left=0.35,
    right=0.95
)

plt.tight_layout()

figure_file = FIGURES_FOLDER / "Figure03_PAS_Top10_Bottom10.png"

plt.savefig(
    figure_file,
    dpi=300,
    bbox_inches="tight"
)

plt.close()
print()

print("✓ Figure 3 saved")

print(figure_file)

# =========================================================
# TABLE 2
# MANUAL LISTENING VS PAS
# =========================================================

manual_poor = {

    "Mabele_Road": "Apple Maps omitted the vowel /e/ ('Mable')",

    "Mqansi_Road": "Incorrect pronunciation identified during listening",

    "Msobho_Road": "Incorrect syllable production",

    "Ndongeni_Road": "Pronunciation differs from native speakers",

    "Ngqwele_Road": "Correctly identified as poor by PAS",

    "Nkathazo_Road": "Stress and phoneme errors",

    "Ntunjambili_Street": "Lowest PAS in dataset",

    "Phezukomkhono_Road": "Multiple pronunciation errors",

    "Phila_Ndwandwe_Road": "Human evaluation poorer than PAS",

    "Qhude_Road": "Click consonant incorrectly realised",

    "Sinqandu_Road": "Incorrect syllable sequence",

    "Umsimbithi_Road": "Pronunciation differs from native speakers",

    "Umzimvubu_Road": "Incorrect pronunciation",

    "Zigqaje_Road": "Pronunciation judged poor by listening"

}

table2 = []

for street, pas in street_pas.items():

    manual = "Poor" if street in manual_poor else "Acceptable"

    if pas >= 0.60:
        pas_rating = "Good"

    elif pas >= 0.45:
        pas_rating = "Moderate"

    else:
        pas_rating = "Poor"

    comparison = (

        "Consistent"

        if manual == pas_rating

        else "Mismatch"

    )

    observation = manual_poor.get(

        street,

        "Pronunciation considered acceptable"

    )

    table2.append({

        "Street Name":

            street.replace("_", " "),

        "Manual Listening":

            manual,

        "PAS (%)":

            round(pas * 100,1),

        "PAS Rating":

            pas_rating,

        "Comparison":

            comparison,

        "Observation":

            observation

    })

table2 = pd.DataFrame(table2)

table2.to_excel(

    "results/Table02_Manual_vs_PAS.xlsx",

    index=False

)

print()

print("✓ Table 2 saved")
table2_file = RESULTS_FOLDER / "Table02_Manual_vs_PAS.xlsx"

table2.to_excel(
    table2_file,
    index=False
)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook(table2_file)
ws = wb.active

# =========================================================
# HEADER STYLE
# =========================================================

header_fill = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

for cell in ws[1]:

    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# =========================================================
# COLOUR COMPARISON COLUMN
# =========================================================

green_fill = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE"
)

red_fill = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE"
)

comparison_column = 5

for row in range(2, ws.max_row + 1):

    cell = ws.cell(row=row, column=comparison_column)

    if cell.value == "Consistent":

        cell.fill = green_fill

    elif cell.value == "Mismatch":

        cell.fill = red_fill

# =========================================================
# AUTO WIDTH
# =========================================================

for column_cells in ws.columns:

    length = max(

        len(str(cell.value)) if cell.value else 0

        for cell in column_cells

    )

    ws.column_dimensions[
        column_cells[0].column_letter
    ].width = length + 3

wb.save(table2_file)

print()
print("✓ Table 2 formatted successfully.")

# =========================================================
# FIGURE 4
# STREET-LEVEL PAS VS DTW
# =========================================================

import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

# ---------------------------------------------------------
# CREATE STREET SUMMARY
# ---------------------------------------------------------

street_summary = (

    results

    .groupby("Street")

    .agg({

        "PAS": "mean",

        "DTW Distance": "mean",

        "Ground Truth": "count"

    })

    .reset_index()

)

street_summary.rename(

    columns={

        "Ground Truth": "Num Recordings"

    },

    inplace=True

)

# ---------------------------------------------------------
# COLOURS
# ---------------------------------------------------------

point_colours = []

for pas in street_summary["PAS"]:

    if pas >= 0.60:

        point_colours.append("#2ca02c")      # Green

    elif pas >= 0.45:

        point_colours.append("#ffbf00")      # Yellow

    else:

        point_colours.append("#d62728")      # Red

# ---------------------------------------------------------
# POINT SIZE
# ---------------------------------------------------------

sizes = street_summary["Num Recordings"] * 40

# ---------------------------------------------------------
# FIGURE
# ---------------------------------------------------------

fig, ax = plt.subplots(figsize=(11,8))

ax.scatter(

    street_summary["DTW Distance"],

    street_summary["PAS"] * 100,

    s=sizes,

    c=point_colours,

    edgecolor="black",

    alpha=0.85

)

# ---------------------------------------------------------
# LABEL EVERY STREET
# ---------------------------------------------------------

for _, row in street_summary.iterrows():

    label = (

        row["Street"]

        .replace("_Road","")

        .replace("_Street","")

        .replace("_Avenue","")

        .replace("_Lane","")

        .replace("_Crescent","")

        .replace("_"," ")

    )

    ax.text(

        row["DTW Distance"] + 120,

        row["PAS"] * 100,

        label,

        fontsize=8

    )

# ---------------------------------------------------------
# MEAN LINES
# ---------------------------------------------------------

mean_dtw = street_summary["DTW Distance"].mean()

mean_pas = street_summary["PAS"].mean()*100

ax.axvline(

    mean_dtw,

    color="red",

    linestyle="--",

    linewidth=1.5

)

ax.axhline(

    mean_pas,

    color="green",

    linestyle="--",

    linewidth=1.5

)

# ---------------------------------------------------------
# TITLES
# ---------------------------------------------------------

ax.set_title(

    "Figure 4. Mean Pronunciation Accuracy Score (PAS)\n"
    "versus Mean DTW Distance for Each Street",

    fontsize=15,

    fontweight="bold",

    pad=20

)

ax.set_xlabel(

    "Mean Dynamic Time Warping (DTW) Distance",

    fontsize=12

)

ax.set_ylabel(

    "Mean Pronunciation Accuracy Score (%)",

    fontsize=12

)

# ---------------------------------------------------------
# GRID
# ---------------------------------------------------------

ax.grid(

    linestyle="--",

    alpha=0.30

)

# ---------------------------------------------------------
# LEGEND
# ---------------------------------------------------------

legend_items = [

    Line2D(

        [0],

        [0],

        marker="o",

        color="w",

        markerfacecolor="#2ca02c",

        markeredgecolor="black",

        markersize=8,

        label="High PAS"

    ),

    Line2D(

        [0],

        [0],

        marker="o",

        color="w",

        markerfacecolor="#ffbf00",

        markeredgecolor="black",

        markersize=8,

        label="Moderate PAS"

    ),

    Line2D(

        [0],

        [0],

        marker="o",

        color="w",

        markerfacecolor="#d62728",

        markeredgecolor="black",

        markersize=8,

        label="Poor PAS"

    )

]

ax.legend(

    handles=legend_items,

    loc="best"

)

plt.tight_layout()

figure_file = FIGURES_FOLDER / "Figure04_StreetLevel_PAS_vs_DTW.png"

plt.savefig(

    figure_file,

    dpi=300,

    bbox_inches="tight"

)

plt.close()

print()

print("✓ Figure 4 saved")

print(figure_file)

# =========================================================
# FIGURE 5
# PITCH COMPARISON
# =========================================================

import matplotlib.pyplot as plt

system_pitch = results["System Pitch"]

ground_truth_pitch = results["Ground Truth Pitch"]

fig, ax = plt.subplots(figsize=(7,6))

box = ax.boxplot(

    [

        system_pitch,

        ground_truth_pitch

    ],

    patch_artist=True,

    labels=[

        "Apple Maps",

        "Native Speakers"

    ]

)

# ---------------------------------------------------------
# Colour boxes
# ---------------------------------------------------------

box["boxes"][0].set_facecolor("#4C78A8")
box["boxes"][1].set_facecolor("#72B7B2")

for whisker in box["whiskers"]:

    whisker.set_color("black")

for cap in box["caps"]:

    cap.set_color("black")

for median in box["medians"]:

    median.set_color("red")

# ---------------------------------------------------------
# Titles
# ---------------------------------------------------------

ax.set_title(

    "Figure 5. Comparison of Pitch Between\n"
    "Apple Maps and Native Speakers",

    fontsize=15,

    fontweight="bold",

    pad=20

)

ax.set_ylabel(

    "Fundamental Frequency (Hz)",

    fontsize=12

)

ax.grid(

    linestyle="--",

    alpha=0.30

)

plt.tight_layout()

figure_file = FIGURES_FOLDER / "Figure05_Pitch_Comparison.png"

plt.savefig(

    figure_file,

    dpi=300,

    bbox_inches="tight"

)

plt.close()

print()

print("✓ Figure 5 saved")

print(figure_file)

# =========================================================
# FIGURE 6
# DURATION COMPARISON
# =========================================================

import matplotlib.pyplot as plt

system_duration = results["System Duration"]

ground_truth_duration = results["Ground Truth Duration"]

fig, ax = plt.subplots(figsize=(7,6))

box = ax.boxplot(

    [

        system_duration,

        ground_truth_duration

    ],

    patch_artist=True,

    tick_labels=[

        "Apple Maps",

        "Native Speakers"

    ]

)

# ---------------------------------------------------------
# Colour boxes
# ---------------------------------------------------------

box["boxes"][0].set_facecolor("#4C78A8")
box["boxes"][1].set_facecolor("#72B7B2")

for whisker in box["whiskers"]:
    whisker.set_color("black")

for cap in box["caps"]:
    cap.set_color("black")

for median in box["medians"]:
    median.set_color("red")

# ---------------------------------------------------------
# Titles
# ---------------------------------------------------------

ax.set_title(

    "Figure 6. Comparison of Pronunciation Duration\n"
    "Between Apple Maps and Native Speakers",

    fontsize=15,

    fontweight="bold",

    pad=20

)

ax.set_ylabel(

    "Duration (Seconds)",

    fontsize=12

)

ax.grid(

    linestyle="--",

    alpha=0.30

)

plt.tight_layout()

figure_file = FIGURES_FOLDER / "Figure06_Duration_Comparison.png"

plt.savefig(

    figure_file,

    dpi=300,

    bbox_inches="tight"

)

plt.close()

print()

print("✓ Figure 6 saved")

print(figure_file)

# =========================================================
# FIGURE 7
# ENERGY COMPARISON
# =========================================================

import matplotlib.pyplot as plt

system_energy = results["System Energy"]

ground_truth_energy = results["Ground Truth Energy"]

fig, ax = plt.subplots(figsize=(7,6))

box = ax.boxplot(

    [

        system_energy,

        ground_truth_energy

    ],

    patch_artist=True,

    tick_labels=[

        "Apple Maps",

        "Native Speakers"

    ]

)

# ---------------------------------------------------------
# Colour boxes
# ---------------------------------------------------------

box["boxes"][0].set_facecolor("#4C78A8")
box["boxes"][1].set_facecolor("#72B7B2")

for whisker in box["whiskers"]:
    whisker.set_color("black")

for cap in box["caps"]:
    cap.set_color("black")

for median in box["medians"]:
    median.set_color("red")

# ---------------------------------------------------------
# Titles
# ---------------------------------------------------------

ax.set_title(

    "Figure 7. Comparison of Speech Energy\n"
    "Between Apple Maps and Native Speakers",

    fontsize=15,

    fontweight="bold",

    pad=20

)

ax.set_ylabel(

    "Average Signal Energy",

    fontsize=12

)

ax.grid(

    linestyle="--",

    alpha=0.30

)

plt.tight_layout()

figure_file = FIGURES_FOLDER / "Figure07_Energy_Comparison.png"

plt.savefig(

    figure_file,

    dpi=300,

    bbox_inches="tight"

)

plt.close()

print()

print("✓ Figure 7 saved")

print(figure_file)
