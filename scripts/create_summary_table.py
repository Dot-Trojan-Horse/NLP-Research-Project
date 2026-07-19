# =========================================================
# CREATE SUMMARY TABLE
# Table 3 - Summary of Experimental Results
# =========================================================

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# =========================================================
# PATHS
# =========================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

RESULTS_FOLDER = PROJECT_ROOT / "results"

INPUT_FILE = RESULTS_FOLDER / "master_results_with_pas.xlsx"

OUTPUT_FILE = RESULTS_FOLDER / "Table03_Experimental_Summary.xlsx"

# =========================================================
# LOAD RESULTS
# =========================================================

print()

print("Loading master results...")

results = pd.read_excel(INPUT_FILE)

print("Done.")

# =========================================================
# CALCULATE SUMMARY
# =========================================================

summary = pd.DataFrame({

    "Metric": [

        "Total street names analysed",

        "Total native speaker recordings analysed",
       
        "Mean Pronunciation Accuracy Score (PAS)",

        "Highest PAS",

        "Lowest PAS",

        "Mean DTW Distance",

        "Mean Normalized DTW",

        "Mean Similarity",

        "Mean Apple Maps Duration (s)",

        "Mean Native Duration (s)",

        "Mean Apple Maps Pitch (Hz)",

        "Mean Native Pitch (Hz)",

        "Mean Apple Maps Energy",

        "Mean Native Energy"

    ],

    "Result": [

        results["Street"].nunique(),

        len(results),

        f"{results['PAS'].mean()*100:.1f}%",

        f"{results['PAS'].max()*100:.1f}%",

        f"{results['PAS'].min()*100:.1f}%",

        f"{results['DTW Distance'].mean():.2f}",

        f"{results['Normalized DTW'].mean():.3f}",
        
        f"{results['Similarity'].mean():.3f}",

        f"{results['System Duration'].mean():.2f}",

        f"{results['Ground Truth Duration'].mean():.2f}",

        f"{results['System Pitch'].mean():.2f}",

        f"{results['Ground Truth Pitch'].mean():.2f}",

        f"{results['System Energy'].mean():.4f}",

        f"{results['Ground Truth Energy'].mean():.4f}"

    ]

})

# =========================================================
# SAVE
# =========================================================

summary.to_excel(

    OUTPUT_FILE,

    index=False

)

print()

print("Summary table saved.")

# =========================================================
# FORMAT EXCEL
# =========================================================

wb = load_workbook(OUTPUT_FILE)

ws = wb.active

header_fill = PatternFill(

    fill_type="solid",

    start_color="1F4E78"

)

header_font = Font(

    bold=True,

    color="FFFFFF"

)

thin = Side(

    style="thin",

    color="BFBFBF"

)

for cell in ws[1]:

    cell.fill = header_fill

    cell.font = header_font

    cell.alignment = Alignment(

        horizontal="center"

    )

for row in ws.iter_rows(min_row=2):

    for cell in row:

        cell.border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin

        )

        cell.alignment = Alignment(

            horizontal="left"

        )

for column in ws.columns:

    length = max(

        len(str(cell.value))

        if cell.value is not None else 0

        for cell in column

    )

    ws.column_dimensions[column[0].column_letter].width = length + 4

wb.save(OUTPUT_FILE)

print()

print("Excel formatting complete.")

print()

print("✓ Table 3 created")

print(OUTPUT_FILE)

